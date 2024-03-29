import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import load_workbook, styles
from fuzzywuzzy import process
import shutil
# Cache for fuzzy matching to avoid redundant computation
fuzzy_match_cache = {}

def delete_specific_rows(sheet):
    rows_to_delete = [row_idx for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1)
                      if all(cell is None for cell in row) or "TOTAL" in (cell for cell in row if cell)]
    for row_idx in reversed(rows_to_delete):
        sheet.delete_rows(row_idx)

def format_sheet(sheet):
    for col in sheet.columns:
        max_length = 0
        for cell in col:
            cell.font = styles.Font(name='Times New Roman', size=10)
            if cell.value and isinstance(cell.value, str):
                max_length = max(max_length, len(cell.value))
        adjusted_width = max(15, (max_length + 2) * 1.1)
        sheet.column_dimensions[col[0].column_letter].width = adjusted_width if col[0].value in ["Description", "Descriptions"] else 15
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 15

def check_for_good_words(sheet, good_words):
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if any(cell for cell in row if cell in good_words):
            return True, row_idx
    return False, None

def delete_rows_above_header(sheet, header_row):
    if header_row > 1:
        sheet.delete_rows(1, header_row - 1)

def rename_columns_based_on_fuzzy_similarity(sheet, good_words):
    for col in sheet.iter_cols(min_row=1, max_row=1, values_only=False):
        col_name = col[0].value
        if col_name and col_name not in fuzzy_match_cache:
            best_match, score = process.extractOne(col_name, good_words)
            if score >= 90:
                fuzzy_match_cache[col_name] = best_match
            else:
                fuzzy_match_cache[col_name] = None
        col[0].value = fuzzy_match_cache.get(col_name, col[0].value)

def delete_columns_not_in_good_words(sheet, good_words):
    columns_to_delete = [col[0].column for col in sheet.iter_cols(min_row=1, max_row=1, values_only=False) if col[0].value not in good_words]
    for col_index in sorted(columns_to_delete, reverse=True):
        sheet.delete_cols(col_index)

def add_additional_columns(sheet, file_name):
    last_col = sheet.max_column + 1
    sheet.cell(row=1, column=last_col, value="GW per unit")
    solo_value = file_name.split("_")[0]
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=last_col + 1, value=solo_value)
    sheet.cell(row=1, column=last_col + 1, value="So lo")

def unmerge_and_process_sheets(workbook, good_words, file_name):
    bad_sheets = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for merge_range in list(sheet.merged_cells.ranges):
            sheet.unmerge_cells(str(merge_range))
        good_sheet, header_row = check_for_good_words(sheet, good_words)
        if good_sheet:
            delete_rows_above_header(sheet, header_row)
            delete_specific_rows(sheet)
            rename_columns_based_on_fuzzy_similarity(sheet, good_words)
            delete_columns_not_in_good_words(sheet, good_words)
        else:
            bad_sheets.append(sheet_name)
    for sheet_name in bad_sheets:
        del workbook[sheet_name]
    if workbook.sheetnames:
        for sheet_name in workbook.sheetnames:
            add_additional_columns(workbook[sheet_name], file_name)
            format_sheet(workbook[sheet_name])
    return workbook.sheetnames

def process_folder(folder_selected, good_words):
    formatted_folder_path = os.path.join(folder_selected, "Formatted")
    bad_files_folder_path = os.path.join(folder_selected, "Bad files")
    os.makedirs(formatted_folder_path, exist_ok=True)
    os.makedirs(bad_files_folder_path, exist_ok=True)

    for file_name in os.listdir(folder_selected):
        if file_name.endswith(".xlsx"):
            full_path = os.path.join(folder_selected, file_name)
            workbook = load_workbook(full_path)
            if unmerge_and_process_sheets(workbook, good_words, file_name):
                # If the file has at least one good sheet, save the processed workbook
                formatted_file_path = os.path.join(formatted_folder_path, file_name)
                workbook.save(formatted_file_path)
            else:
                # If all sheets are bad, copy the file to the "Bad files" folder
                bad_file_path = os.path.join(bad_files_folder_path, file_name)
                shutil.copy2(full_path, bad_file_path)

def main():
    root = tk.Tk()
    root.withdraw()  # Hide the Tkinter root window
    folder_selected = filedialog.askdirectory()  # Prompt the user to select a directory
    if folder_selected:
        good_words = ["SKU", "Description", "Descriptions", "Material", "Quantity", "Qty", "G/W(KG)", "Gross Weight", "SKU No."]
        process_folder(folder_selected, good_words)

if __name__ == "__main__":
    main()

